from fillpdf import fillpdfs
from openpyxl import load_workbook
import copy
import re
from dateutil.parser import parse
from datetime import datetime
import PySimpleGUI as sg
import os
import sys
import pkg_resources
from pdf2jpg import pdf2jpg

# Define GUI layout
layout = [
    [sg.Text("Input File"), sg.Input(key="-IN-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Text("Output Folder"), sg.Input(key="-OUT-"), sg.FolderBrowse()],
    [sg.Text("Form Folder Name"), sg.Input(key="-FORM-")],
    [sg.Button("Fill Forms")],
    [sg.Text("1. Click 'Browse' next to 'Input File' to select your Excel (.xlsx) file of choice.")],
    [sg.Text("2. Click 'Browse' next to 'Output Folder' to select the destination of the generated folder.")],
    [sg.Text("3. Name the folder in which the filled forms will be saved.")],
    [sg.Text("*Note: Program will overwrite folders with the same name without asking for confirmation.")],
    [sg.Text("4. Click 'Fill Forms'")]
]

# Create the GUI window
window = sg.Window("CLE Form Filler", layout)

# Event loop to process GUI events
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED:
        break

    if event == "Fill Forms":
        # Get user input values
        input_file = values["-IN-"]
        output_folder = values["-OUT-"]
        form_folder = values["-FORM-"]

        # Specify output folder
        output_folder = os.path.join(output_folder, form_folder)
        os.makedirs(output_folder, exist_ok=True)

        # Perform PDF form filling
        data = fillpdfs.get_form_fields('CLE_V3.pdf')
        dataframe = load_workbook(input_file)
        sheet = dataframe.active
        max_rows = sheet.max_row

        for row_num in range(2, max_rows + 1):
            # Get the current row
            current_row = sheet[row_num]

            # Check if the first column (column 'A') is empty
            if current_row[0].value is None:
                break  # Exit loop if row is empty

             # Section A
            email = current_row[0].value
            Last_Name = current_row[1].value
            First_Name = current_row[2].value
            Program_Title = current_row[3].value

            ###################

            # Define the maximum number of words per line
            max_words_per_line = 3

            # Split the Program_Title into words
            words = Program_Title.split()

            # Create a list to store the lines
            lines = []

            # Iterate over the words and split them into lines
            current_line = ""
            for word in words:
                current_line += word + " "

                # Check if the current line exceeds the maximum words per line
                if len(current_line.split()) >= max_words_per_line:
                    lines.append(current_line.strip())
                    current_line = ""

            # Add the remaining words as the last line
            if current_line:
                lines.append(current_line.strip())

            # Join the lines with line breaks
            Program_Title = "\n".join(lines)


            ###################

            # Work on time format
            #Attend_Date = current_row[4].value
            #if isinstance(Attend_Date, (int, float)):
            #    Attend_Date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + Attend_Date - 2)
            #    Attend_Date = Attend_Date.strftime('%m/%d/%Y')

            # Work on time format
            Attend_Date = current_row[4].value
            if isinstance(Attend_Date, datetime):
                Attend_Date = Attend_Date.strftime('%m/%d/%Y')

            #function to make a cell of type None to an empty string
            def make_empty_string(attribute):
                if attribute == None:
                    attribute = ""
                return attribute

             # Section B
            Ethics = current_row[5].value
            Ethics = make_empty_string(Ethics)

            Skills = current_row[6].value
            Skills = make_empty_string(Skills)

            Law_Practice = current_row[7].value
            Law_Practice = make_empty_string(Law_Practice)

            Area_of_Practice = current_row[8].value
            Area_of_Practice = make_empty_string(Area_of_Practice)

            CPD_Ethics = current_row[9].value
            CPD_Ethics = make_empty_string(CPD_Ethics)

            CPD_General = current_row[10].value
            CPD_General = make_empty_string(CPD_General)

            Bias = current_row[11].value
            Bias = make_empty_string(Bias)

            # Section C
            Credit_Faculty_Part = current_row[12].value
            # conditional needed for values
            if Credit_Faculty_Part == "Speaker":
                Credit_Faculty_Part = 0
            elif Credit_Faculty_Part == "Moderator":
                Credit_Faculty_Part = 1
            elif Credit_Faculty_Part == "Panel":
                Credit_Faculty_Part = 2
            elif Credit_Faculty_Part == "LCF":
                Credit_Faculty_Part = 3
            else:
                Credit_Faculty_Part = None

            Ethics_Prof_Part = current_row[13].value
            Ethics_Prof_Part = make_empty_string(Ethics_Prof_Part)

            Skills_Part = current_row[14].value
            Skills_Part = make_empty_string(Skills_Part)

            Law_Practice_Part = current_row[15].value
            Law_Practice_Part = make_empty_string(Law_Practice_Part)

            Area_of_Practice_Part = current_row[16].value
            Area_of_Practice_Part = make_empty_string(Area_of_Practice_Part)

            CPD_Ethics_Part = current_row[17].value
            CPD_Ethics_Part = make_empty_string(CPD_Ethics_Part)

            CPD_General_Part = current_row[18].value
            CPD_General_Part = make_empty_string(CPD_General_Part)

            Bias_Part = current_row[19].value
            Bias_Part = make_empty_string(Bias_Part)

            # Section D
            Class_Format = current_row[20].value
            Class_Text = copy.copy(Class_Format)
            Class_Format = Class_Format.lower()

            # conditional needed for values
            if Class_Format == "1 traditional":
                Class_Format = 0
                Class_Text = ""
            elif Class_Format == "2 lst with questions":
                Class_Format = 1
                Class_Text = ""
            elif Class_Format == "3 lst without questions":
                Class_Format = 2
                Class_Text = ""
            elif Class_Format == "4 fully interactive":
                Class_Format = 3
                Class_Text = ""
            elif Class_Format == "5 Prerecorded":
                Class_Format = 4
                Class_Text = ""
            else:
                Class_Format = 5
             # Section E
            Method_Part = current_row[21].value
            Method_Part = Method_Part.lower()
            # conditional needed for values
            if Method_Part == "individual":
                Method_Part = 0
            elif Method_Part == "group":
                Method_Part = 1
             # Section F
            Course_Cont = current_row[22].value
            Course_Cont = Course_Cont.lower()
            # conditional needed for values
            if Course_Cont == "both":
                Course_Cont = 0
            # CASE SENSITIVITY RESOLVED
            elif Course_Cont == "only experienced":
                Course_Cont = 1
            elif Course_Cont == "only newly":
                Course_Cont = 2

            
             # Section G
            Provider_Org = "QUEENS COUNTY BAR\nASSOCIATION"
            CLE_Provider = current_row[29].value
            CLE_Provider = CLE_Provider.lower()
            try:
                if CLE_Provider == "yes":
                    CLE_Provider = 0
            except:
                print("Check Certification")

            #Agent_Name = "Jonathan Riegel, Exec. Director"
            # print(cell.value)
            # Dictionary of key value pairs
            # Keys are the PDF field names
            # Values are the values of the XLSX cells
            data.update({"Text Field A1": Last_Name + ', ' + First_Name,
                        "Text Field A2": Program_Title,
                         "Text Field A3": Attend_Date,
                         "Text Field B1": Ethics,
                         "Text Field B2": Skills,
                         "Text Field B3": Law_Practice,
                         "Text Field B4": Area_of_Practice,
                         "Text Field B5": CPD_Ethics,
                         "Text Field B6": CPD_General,
                         "Text Field B7": Bias,
                         "Radio Button C": Credit_Faculty_Part,
                         "Text Field C1": Ethics_Prof_Part,
                         "Text Field C2": Skills_Part,
                         "Text Field C3": Law_Practice_Part,
                         "Text Field C4": Area_of_Practice_Part,
                         "Text Field C5": CPD_Ethics_Part,
                         "Text Field C6": CPD_General_Part,
                         "Text Field C7": Bias_Part,
                         "Radio Button D": Class_Format,
                         "Text Field D1": Class_Text,
                         "Radio Button E": Method_Part,
                         "Radio Button F": Course_Cont,
                         "Text Field G1": Provider_Org,
                         "Radio Button G": CLE_Provider,
                         #"Text Field G5": Agent_Name
                         })

            print(data["Text Field A1"])
            # Specify output filename
            output_filename = os.path.join(output_folder, Last_Name + '_' + First_Name + '.pdf')
            # Perform PDF form filling
            fillpdfs.write_fillable_pdf('CLE_V3.pdf', output_filename, data, flatten=False)
            #Flatten via image method to help with cross-compatability
            fillpdfs.flatten_pdf(output_filename, output_filename, as_images=True)
            
    print("Filling Complete!\nYou may close the program and navigate to your folder.")
    sg.popup(f"Filling Complete!\nYou may close the program and navigate to your folder.", title="CLE Form Filler")
    window.close()           